import streamlit as st
import pandas as pd
import openai
from fuzzywuzzy import fuzz
import openpyxl
from openpyxl.styles import PatternFill, Font
import matplotlib.pyplot as plt
import seaborn as sns

# OpenAI API Key
openai.api_key = "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"
st.image("https://strapi.tarento.com/uploads/Tarento_logo_749f934596.svg", width=200)

# Function to compute LLM-based semantic similarity
def compute_llm_similarity(input_value, reference_value):
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": "You are an assistant evaluating semantic similarity between two texts."},
                {"role": "user",
                 "content": f"Rate the similarity between these two values on a scale of 0 to 1:\n\nValue 1: {input_value}\nValue 2: {reference_value}"}
            ],
        )
        score = float(response["choices"][0]["message"]["content"].strip())
        return score
    except Exception as e:
        print(f"Error during LLM similarity evaluation: {e}")
        return 0  # Default to 0 similarity in case of error

# Function to preprocess strings
def preprocess_string(value):
    return value.strip().lower().replace(" ", "")

# Streamlit UI
st.title("Semantic Similarity Scoring Tool")
st.sidebar.header("Settings")

uploaded_ref_file = st.sidebar.file_uploader("Upload Reference File", type=["xlsx"])

if uploaded_ref_file:
    try:
        ref_df = pd.read_excel(uploaded_ref_file, dtype=str)
        st.session_state["ref_columns"] = ref_df.columns.tolist()
        st.sidebar.success("Reference file loaded successfully!")
    except Exception as e:
        st.sidebar.error(f"Error loading reference file: {e}")
else:
    st.sidebar.warning("Please upload a reference file.")

if "ref_columns" not in st.session_state:
    st.session_state["ref_columns"] = []

mandatory_columns = st.sidebar.multiselect("Select Mandatory Columns", st.session_state["ref_columns"], key="mandatory")
good_to_have_columns = st.sidebar.multiselect("Select Good-to-Have Columns", st.session_state["ref_columns"], key="good_to_have")

uploaded_input_file = st.sidebar.file_uploader("Upload Input File", type=["xlsx"])
# Add weightage selection in the sidebar
st.sidebar.subheader("Weightage Settings")
weight_option = st.sidebar.radio(
    "Select weightage between Mandatory and Good-to-Have fields:",
    options=["70:30", "60:40"],
    index=0  # Default to the first option
)

# Adjust weights based on user selection
if weight_option == "70:30":
    mandatory_weight = 70
    good_to_have_weight = 30
else:
    mandatory_weight = 60
    good_to_have_weight = 40

st.sidebar.write(f"Selected Weightage: {mandatory_weight}% Mandatory, {good_to_have_weight}% Good-to-Have")
output_path = st.sidebar.text_input("Output File Path", "output.xlsx")

if st.sidebar.button("Run Scoring"):
    with st.spinner("Processing data... Please wait."):
        if uploaded_input_file and uploaded_ref_file:
            try:
                # Load files
                input_df = pd.read_excel(uploaded_input_file, dtype=str)
                ref_df = pd.read_excel(uploaded_ref_file, dtype=str)

                results = []
                updated_input = input_df.copy()  # Create a copy of the input file for updates
                # Initialize overall counters
                matched_mandatory_count = 0
                partial_mandatory_count = 0
                non_matched_mandatory_count = 0

                matched_good_to_have_count = 0
                partial_good_to_have_count = 0
                non_matched_good_to_have_count = 0

                # Initialize per-column counters
                fully_matched_counts = {col: 0 for col in mandatory_columns + good_to_have_columns}
                partial_matched_counts = {col: 0 for col in mandatory_columns + good_to_have_columns}
                non_matched_counts = {col: 0 for col in mandatory_columns + good_to_have_columns}

                for _, row in ref_df.iterrows():
                    row_score = 0
                    mandatory_score = 0
                    good_to_have_score = 0

                    fully_matched_columns = []
                    fully_matched_values = []
                    partially_matched_columns = []
                    partially_matched_values = []

                    mandatory_values = []
                    good_to_have_values = []
                    mandatory_match_found = False

                    # Process mandatory and good-to-have columns
                    for col, weight in [(mandatory_columns, mandatory_weight), (good_to_have_columns, good_to_have_weight)]:
                        for input_col in col:
                            input_value = str(row.get(input_col, "")).strip() if pd.notna(row.get(input_col)) else ""

                            if input_col in mandatory_columns:
                                mandatory_values.append(input_value)
                            if input_col in good_to_have_columns:
                                good_to_have_values.append(input_value)

                            if not input_value or input_value == "No Data":
                                non_matched_counts[input_col] += 1
                                continue

                            full_match_found = False
                            max_similarity = 0
                            best_ref_value = None
                            best_ref_col = None

                            # Compare input value against all reference columns
                            for ref_col in input_df.columns:
                                ref_values = input_df[ref_col].dropna().astype(str).tolist()

                                # Check for full match
                                if input_value in ref_values:
                                    row_score += weight
                                    if input_col in mandatory_columns:
                                        mandatory_score += weight
                                        matched_mandatory_count += 1
                                        mandatory_match_found = True
                                    else:
                                        good_to_have_score += weight
                                        matched_good_to_have_count += 1

                                    fully_matched_counts[input_col] += 1
                                    fully_matched_columns.append(f"{input_col} -> {ref_col}")
                                    fully_matched_values.append(input_value)
                                    full_match_found = True
                                    break

                                # Skip partial matching if full match found
                                if full_match_found:
                                    continue

                                # Use LLM for partial matching
                                for ref_value in ref_values:
                                    if ref_value in fully_matched_values:
                                        continue

                                    processed_input = preprocess_string(input_value)
                                    processed_ref = preprocess_string(ref_value)

                                    llm_similarity = compute_llm_similarity(processed_input, processed_ref)
                                    if llm_similarity > max_similarity:
                                        max_similarity = llm_similarity
                                        best_ref_value = ref_value
                                        best_ref_col = ref_col

                            # Handle partial matches
                            if max_similarity > 0.3:
                                row_score += max_similarity * weight
                                if input_col in mandatory_columns:
                                    mandatory_score += max_similarity * weight
                                    partial_mandatory_count += 1
                                    mandatory_match_found = True
                                else:
                                    good_to_have_score += max_similarity * weight
                                    partial_good_to_have_count += 1

                                partial_matched_counts[input_col] += 1
                                partially_matched_columns.append(f"{input_col} -> {best_ref_col}")
                                partially_matched_values.append(best_ref_value)
                            elif not full_match_found:
                                non_matched_counts[input_col] += 1

                    # If any mandatory column matches or partially matches, add all ref_df columns
                    if mandatory_match_found:
                        for col in ref_df.columns:
                            if col not in updated_input.columns:
                                updated_input[col] = ref_df[col]

                    # Normalize scores
                    max_possible_score = (len(mandatory_columns) * mandatory_weight) + (len(good_to_have_columns) * good_to_have_weight)
                    normalized_score = (row_score / max_possible_score) * 100
                    mandatory_score_normalized = (mandatory_score / (len(mandatory_columns) * mandatory_weight)) * 100
                    good_to_have_score_normalized = (good_to_have_score / (len(good_to_have_columns) * good_to_have_weight)) * 100

                    # Determine priority
                    priority = 1 if normalized_score > 90 else 2 if 75 <= normalized_score <= 90 else 3 if 50 <= normalized_score < 75 else 4 if 10 <= normalized_score < 50 else 5

                    results.append({

                        "Mandatory_Columns": ", ".join(mandatory_columns),
                        "Good_to_Have_Columns": ", ".join(good_to_have_columns),
                        "Mandatory_Columns_Values": ", ".join(mandatory_values),
                        "Good_to_Have_Columns_Values": ", ".join(good_to_have_values),
                        "Fully_Matched_Columns": ", ".join(fully_matched_columns),
                        "Fully_Matched_Values": ", ".join(fully_matched_values),
                        "Partially_Matched_Columns": ", ".join(partially_matched_columns),
                        "Partially_Matched_Values": ", ".join(partially_matched_values),
                        "Normalized_Score": round(normalized_score, 2),
                        "Priority": priority,
                        "Mandatory_Score": round(mandatory_score_normalized, 2),
                        "Good_to_Have_Score": round(good_to_have_score_normalized, 2),
                        "ID_S": row.get("ID_S", "No Data")
                    })

                # Add new columns to input_df
                #   if "Manufacturer_S" in ref_df.columns and "Manufacturers_Item_S" in ref_df.columns:
                #      input_df["Manufacturer_S"] = ref_df["Manufacturer_S"]
                #     input_df["Manufacturers_Item_S"] = ref_df["Manufacturers_Item_S"]

                # Categorize rows into Fully Matched, Partially Matched, and Not Matched
                fully_matched_rows = [row for row in results if row["Priority"] == 1]
                partially_matched_rows = [row for row in results if row["Priority"] in [2, 3, 4]]
                not_matched_rows = [row for row in results if row["Priority"] == 5]

                # Add recommendations for "Not Matched" rows
                for row in not_matched_rows:
                    unmatched_cols = row["Mandatory_Columns"] + row["Good_to_Have_Columns"]
                    recommendations = []
                    for col in unmatched_cols.split(", "):
                        input_value = str(row.get(col, ""))
                        if input_value:
                            best_match = None
                            max_similarity = 0
                            for ref_col in input_df.columns:
                                for ref_value in input_df[ref_col].dropna().astype(str):
                                    similarity = fuzz.ratio(input_value.lower(), ref_value.lower()) / 100.0
                                    if similarity > max_similarity:
                                        max_similarity = similarity
                                        best_match = ref_value
                            recommendations.append(f"{col}: {best_match} (Similarity: {max_similarity:.2f})")
                    row["Recommendations"] = "; ".join(recommendations)

                fully_matched_df = pd.DataFrame(fully_matched_rows)
                partially_matched_df = pd.DataFrame(partially_matched_rows)
                not_matched_df = pd.DataFrame(not_matched_rows)
                # Visualization - Charts for Home Page
                # Visualization - Charts for Home Page
                # Calculate row-level counts
                row_counts = {
                    "Fully Matched": len(fully_matched_rows),
                    "Partially Matched": len(partially_matched_rows),
                    "Not Matched": len(not_matched_rows),
                }

                # Row-Level Matching Overview
                st.markdown("### Row-Level Matching Overview")

                # Add context text to explain the graph
                st.markdown(
                    """
                    This bar chart provides an overview of the row-level matching performance:
    
                    - **Fully Matched**: Rows that perfectly match between datasets.
                    - **Partially Matched**: Rows that have some matching elements but are not a complete match.
                    - **Not Matched**: Rows that do not have any match in the comparison dataset.
                    """
                )

                # Create bar chart for match counts
                fig_match_counts, ax_match_counts = plt.subplots(
                    figsize=(8, 6))  # Adjust the figure size for better visibility
                sns.barplot(
                    x=list(row_counts.keys()),
                    y=list(row_counts.values()),
                    palette="viridis",
                    ax=ax_match_counts
                )

                # Enhance the title and axes labels
                ax_match_counts.set_title(
                    "Row-Level Matching Performance\n(Fully Matched, Partially Matched, and Not Matched Rows)",
                    fontsize=14, weight='bold'
                )
                ax_match_counts.set_xlabel("Row Match Type", fontsize=12, labelpad=10)
                ax_match_counts.set_ylabel("Number of Rows", fontsize=12, labelpad=10)

                # Remove decimals from y-axis
                ax_match_counts.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{int(x):,}"))

                # Add counts on top of bars
                for i, count in enumerate(row_counts.values()):
                    ax_match_counts.text(i, count + max(row_counts.values()) * 0.02, f"{count:,}",
                                         # Add thousands separator
                                         ha='center', va='bottom', fontsize=10, color='black')

                # Add gridlines for better readability
                ax_match_counts.yaxis.grid(True, linestyle='--', alpha=0.7)
                ax_match_counts.set_axisbelow(True)

                # Display the plot in Streamlit
                st.pyplot(fig_match_counts)

                # Add a footer note
                st.markdown(
                    """
                    *Note*: This chart helps in analyzing the distribution of row-level matches, providing insights into data alignment quality.
                    """
                )

                # Create a container to hold both graphs side by side
                with st.container():
                    col1, col2 = st.columns(2)  # Define two equal-width columns

                    # Left Column: Recommendations Effectiveness
                    with col1:
                        st.markdown("### Recommendations Effectiveness")
                        st.markdown(
                            "This section provides insights into the recommendations generated for unmatched records. "
                            "It highlights how many records received recommendations and how many did not."
                        )

                        # Recommendations data
                        total_unmatched_records = len(not_matched_rows)
                        recommendations_provided = sum(
                            1 for row in not_matched_rows if "Recommendations" in row and row["Recommendations"]
                        )
                        no_recommendations = total_unmatched_records - recommendations_provided
                        recommendation_percentage = (
                            (recommendations_provided / total_unmatched_records) * 100
                            if total_unmatched_records > 0 else 0
                        )

                        # Pie Chart Data
                        recommendations_data = {
                            "Category": ["Recommendations Provided", "No Recommendations"],
                            "Count": [recommendations_provided, no_recommendations],
                        }
                        recommendations_df = pd.DataFrame(recommendations_data)

                        # Plot Pie Chart
                        fig_recommendations, ax_recommendations = plt.subplots(figsize=(5, 4))
                        ax_recommendations.pie(
                            recommendations_df["Count"],
                            labels=recommendations_df["Category"],
                            autopct='%1.1f%%',
                            startangle=140,
                            colors=["#32CD32", "#FF6347"],
                        )
                        ax_recommendations.set_title("Recommendations Breakdown")
                        st.pyplot(fig_recommendations)

                        # Summary Details
                        st.markdown(
                            f"**Summary:**\n\n"
                            f"- **Total Unmatched Records:** {total_unmatched_records}\n"
                            f"- **Recommendations Provided:** {recommendations_provided} "
                            f"({recommendation_percentage:.2f}% of unmatched records)\n"
                            f"- **Records Without Recommendations:** {no_recommendations}"
                        )

                    # Right Column: Similarity Scores Distribution
                    with col2:
                        st.markdown("### Similarity Scores Distribution")
                        st.markdown(
                            "This graph shows the distribution of similarity scores for both **mandatory** and **good-to-have** columns, helping you understand how well the records match."
                        )
                        fig_similarity, ax_similarity = plt.subplots(figsize=(5, 4))

                        # Similarity scores for mandatory columns
                        mandatory_similarity_scores = [
                            row["Mandatory_Score"] for row in results if row["Mandatory_Score"] > 0
                        ]
                        sns.histplot(
                            mandatory_similarity_scores, kde=True, bins=15, color='blue',
                            label="Mandatory Columns", alpha=0.6, ax=ax_similarity
                        )

                        # Similarity scores for good-to-have columns
                        good_to_have_similarity_scores = [
                            row["Good_to_Have_Score"] for row in results if row["Good_to_Have_Score"] > 0
                        ]
                        sns.histplot(
                            good_to_have_similarity_scores, kde=True, bins=15, color='green',
                            label="Good-to-Have Columns", alpha=0.6, ax=ax_similarity
                        )

                        # Adding Descriptive Titles and Labels
                        ax_similarity.set_title("Distribution of Similarity Scores", fontsize=14)
                        ax_similarity.set_xlabel("Similarity Score (0 = No Match, 100 = Perfect Match)", fontsize=12)
                        ax_similarity.set_ylabel("Frequency (Number of Records)", fontsize=12)

                        # Add a legend for clarity
                        ax_similarity.legend(title="Field Type", fontsize=10, title_fontsize=12)

                        st.pyplot(fig_similarity)

                # Spacer for separation
                st.markdown("---")

                # Comparison of Full Matches Across Columns
                st.markdown("### Comparison of Full Matches Across Columns")

                # Prepare data for the grouped bar chart
                comparison_data = {
                    "Column": mandatory_columns + good_to_have_columns,
                    "Fully Matched": [fully_matched_counts[col] for col in mandatory_columns + good_to_have_columns],
                    "Partially Matched": [partial_matched_counts[col] for col in mandatory_columns + good_to_have_columns],
                    "Non-Matched": [non_matched_counts[col] for col in mandatory_columns + good_to_have_columns],
                }

                comparison_df = pd.DataFrame(comparison_data)

                # Melt the DataFrame for easier plotting
                melted_df = comparison_df.melt(id_vars="Column", var_name="Match Type", value_name="Count")

                # Create the grouped bar chart
                fig_comparison, ax_comparison = plt.subplots(figsize=(10, 6))
                sns.barplot(
                    x="Column",
                    y="Count",
                    hue="Match Type",
                    data=melted_df,
                    palette="muted",
                    ax=ax_comparison
                )

                # Customize the chart
                ax_comparison.set_title("Comparison of Full Matches, Partial Matches, and Non-Matches Across Columns",
                                        fontsize=14)
                ax_comparison.set_xlabel("Columns", fontsize=12)
                ax_comparison.set_ylabel("Count", fontsize=12)
                ax_comparison.set_xticklabels(ax_comparison.get_xticklabels(), rotation=45, ha="right", fontsize=10)
                ax_comparison.legend(title="Match Type", fontsize=10)

                # Display the chart in Streamlit
                st.pyplot(fig_comparison)

                # Priority Distribution
                # Ranking and Significance Distribution
                st.markdown("### Ranking and Significance Distribution of Records")

                # Explanation for the user
                st.markdown(
                    """
                    This graph visualizes the distribution of records based on their assigned ranking levels, 
                    which indicate their significance or importance. The rankings are defined as follows:
                    - **1 = Very Low Significance**: These records do not require immediate attention or action.
                    - **2 = Low Significance**: Tasks or records that can generally be deferred.
                    - **3 = Moderate Significance**: Standard tasks that may require attention in due course.
                    - **4 = High Significance**: Tasks that should be addressed soon.
                    - **5 = Very High Significance**: Critical records requiring immediate action.
    
                    Use this graph to prioritize your focus on higher-ranking records (Ranks 4 and 5), 
                    while deprioritizing or ignoring those marked as Rank 1.
                    """
                )

                # Extract ranking data
                ranking_data = [row["Priority"] for row in results]  # Assuming "Priority" column represents ranking levels

                # Create the plot
                fig_ranking, ax_ranking = plt.subplots(figsize=(6, 4))
                sns.countplot(x=ranking_data, palette="coolwarm", ax=ax_ranking)

                # Set titles and labels
                ax_ranking.set_title("Ranking and Significance Distribution")
                ax_ranking.set_xlabel("Ranking (1 = Very Low Significance, 5 = Very High Significance)")
                ax_ranking.set_ylabel("Number of Records")

                # Render the plot
                st.pyplot(fig_ranking)

                # AI Explanation of Advantages
                st.markdown("### How AI Enhances Your Experience")
                st.markdown("""
                - **Semantic Matching**: AI compares records semantically, ensuring contextually similar values are linked even if they aren't exact matches.
                - **Recommendations**: For unmatched records, AI suggests possible matches with confidence levels.
                - **Prioritization**: Records are ranked based on their similarity scores, helping you focus on the most critical ones.
                - **Category Analysis**: Separate analysis for Mandatory and Good-to-Have fields helps prioritize effectively.
                - **Insights**: Detailed visualizations give you a clear understanding of matching effectiveness, recommendations, and overall results.
                """)

                # Prepare Fully Matched and Partially Matched DataFrames
                fully_matched_rows = []
                partially_matched_rows = []

                # After matching rows and determining the final output, ensure to add input file columns to these rows.

                # After matching rows and determining the final output, ensure to add input file columns only when mandatory columns match.

                for result in results:
                    priority = result["Priority"]
                    matched_id = result["ID_S"]

                    # Locate the matching row in the reference file
                    ref_row = ref_df[ref_df["ID_S"] == matched_id].to_dict(orient="records")
                    if ref_row:
                        ref_row = ref_row[0]  # Extract the matching row dictionary

                        # Combine the result with reference data (initially without input file columns)
                        combined_row = {**result, **ref_row}

                        # Check if mandatory columns have matched (either fully or partially)
                        mandatory_match_found = False
                        for col in mandatory_columns:
                            if col in fully_matched_columns or col in partially_matched_columns:
                                mandatory_match_found = True
                                break

                        # Add input file columns only if mandatory columns match
                        if mandatory_match_found:
                            for input_col in input_df.columns:
                                combined_row[input_col] = input_df[input_col].iloc[
                                    0]  # Add input columns (from the first row or specific logic)

                        # Add to respective tabs based on priority
                        if priority == 1:  # Fully Matched
                            fully_matched_rows.append(combined_row)
                        elif priority in [2, 3, 4]:  # Partially Matched
                            partially_matched_rows.append(combined_row)

                # Convert Fully Matched and Partially Matched rows to DataFrames
                fully_matched_df = pd.DataFrame(fully_matched_rows)
                partially_matched_df = pd.DataFrame(partially_matched_rows)

                # Ensure all reference columns are included in the final DataFrame
                for col in ref_df.columns:
                    if col not in fully_matched_df.columns:
                        fully_matched_df[col] = None
                    if col not in partially_matched_df.columns:
                        partially_matched_df[col] = None

                # Save to Excel file
                with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                    # Save other output tabs
                    final_df = pd.DataFrame(results)
                    final_df.to_excel(writer, sheet_name="Results", index=False)
                    updated_input.to_excel(writer, sheet_name="Updated Input", index=False)
                    # Write Fully Matched and Partially Matched tabs
                    fully_matched_df.to_excel(writer, sheet_name="Fully Matched", index=False)
                    partially_matched_df.to_excel(writer, sheet_name="Partially Matched", index=False)
                    not_matched_df = pd.DataFrame(not_matched_rows)
                    not_matched_df.to_excel(writer, sheet_name="Not Matched", index=False)

                    # Apply formatting to highlight new columns in Fully Matched
                    workbook = writer.book
                    fully_matched_ws = workbook["Fully Matched"]
                    partially_matched_ws = workbook["Partially Matched"]

                    # Light yellow fill for new columns
                    fill_new_columns = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

                    # Find new columns and highlight them
                    for col in ref_df.columns:
                        if col not in input_df.columns:  # New column added from ref_df
                            col_idx_fully = fully_matched_df.columns.get_loc(col) + 1  # 1-based index
                            col_idx_partial = partially_matched_df.columns.get_loc(col) + 1  # 1-based index

                            # Apply fill to header and entire column in Fully Matched
                            fully_matched_ws.cell(row=1, column=col_idx_fully).fill = fill_new_columns
                            for row in range(2, fully_matched_df.shape[0] + 2):  # Start from row 2 for data
                                fully_matched_ws.cell(row=row, column=col_idx_fully).fill = fill_new_columns

                            # Apply fill to header and entire column in Partially Matched
                            partially_matched_ws.cell(row=1, column=col_idx_partial).fill = fill_new_columns
                            for row in range(2, partially_matched_df.shape[0] + 2):  # Start from row 2 for data
                                partially_matched_ws.cell(row=row, column=col_idx_partial).fill = fill_new_columns

                            # Apply color to new column headers in "Updated Input"
                            workbook = writer.book
                            worksheet = workbook["Updated Input"]
                            fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

                            # Apply fill color to the newly added column headers
                            for col in ref_df.columns:
                                if col in updated_input.columns:
                                    col_idx = updated_input.columns.get_loc(col) + 1
                                    worksheet.cell(row=1, column=col_idx).fill = fill

                st.success(f"Scoring completed. Results saved to {output_path}")

                # Display Mandatory and Good-to-Have Columns Table
                mandatory_counts_df = pd.DataFrame({
                    "Column": [col for col in mandatory_columns],
                    "Fully Matched": [fully_matched_counts[col] for col in mandatory_columns],
                    "Partially Matched": [partial_matched_counts[col] for col in mandatory_columns],
                    "Non-Matched": [non_matched_counts[col] for col in mandatory_columns],
                })

                good_to_have_counts_df = pd.DataFrame({
                    "Column": [col for col in good_to_have_columns],
                    "Fully Matched": [fully_matched_counts[col] for col in good_to_have_columns],
                    "Partially Matched": [partial_matched_counts[col] for col in good_to_have_columns],
                    "Non-Matched": [non_matched_counts[col] for col in good_to_have_columns],
                })

                st.markdown("<h4 style='color: blue;'>Mandatory Columns</h4>", unsafe_allow_html=True)
                st.dataframe(mandatory_counts_df)

                st.markdown("<h4 style='color: green;'>Good-to-Have Columns</h4>", unsafe_allow_html=True)
                st.dataframe(good_to_have_counts_df)
            except Exception as e:
                st.error(f"Error: {e}")
        else:
            st.error("Please upload both input and reference files.")

# Custom CSS for button styling
st.markdown(
    """
    <style>
        .custom-button {
            background-color: #4CAF50; /* Green */
            color: white;
            font-size: 16px;
            font-weight: bold;
            padding: 10px 20px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            transition: 0.3s;
        }
        .custom-button:hover {
            background-color: #45a049; /* Darker Green */
        }
    </style>
    """,
    unsafe_allow_html=True,
)

# Sidebar Help Button with Custom Style
if st.sidebar.markdown('<button class="custom-button">ℹ️ Help Guide</button>', unsafe_allow_html=True):


    # Expandable help guide section
    with st.sidebar.expander("📖 Help Guide - Understanding This App"):
        st.markdown("""
        ### 📌 What This App Does:
        - This tool evaluates **semantic similarity** between two datasets.
        - Uses **AI-powered matching** (GPT-4) and **fuzzy matching** to compare values.
        - Generates an **Excel report** with **Fully Matched, Partially Matched, and Not Matched** data.
        - Provides **visualizations** for better insights.

        ---

        ### 🔹 **UI Components Explained**
        #### 📂 1. Upload Files
        - **Reference File** (Required) → The dataset to compare against.
        - **Input File** (Required) → The dataset to be matched.

        #### 🎛️ 2. Matching Settings
        - **Mandatory Columns** → Fields that must match.
        - **Good-to-Have Columns** → Optional fields that improve the score.
        - **Weightage (70:30 or 60:40)** → How much weight Mandatory vs Good-to-Have fields get.

        #### 📊 3. Visual Reports
        - **Bar Chart**: Overview of fully, partially, and unmatched records.
        - **Pie Chart**: Effectiveness of AI recommendations.
        - **Histogram**: Distribution of similarity scores.

        #### 📝 4. Output File
        - The results are saved in **`output.xlsx`**.
        - Includes separate sheets for **fully matched, partially matched, and not matched data**.

        ---

        ### ❓ How to Use:
        1️⃣ **Upload Reference & Input Files**.  
        2️⃣ **Select Mandatory & Good-to-Have Columns**.  
        3️⃣ **Choose Weightage and Click "Run Scoring"**.  
        4️⃣ **Check Results in the Excel Output**.  
        5️⃣ **Use the Help Guide if needed!**  

        🔹 _Need further assistance? Contact support!_ 💡
        """)
